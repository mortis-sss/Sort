from openpyxl import load_workbook as lw
import time
import random
import matplotlib.pyplot as plt

plt.ion()

file = lw('1.xlsx')
pages = ['03.04.2022']
page = file[pages[0]]
A = []
B = []
C = []

for i in range(2, 200):
    try:
        A.append(float(page.cell(row=i, column=8).value))
        B.append(float(page.cell(row=i, column=9).value))
        C.append(float(page.cell(row=i, column=10).value))
    except:
        var = None


def shellsort(array):
    print('shell')
    a = len(array)
    gap = a // 2
    while gap > 0:
        graf(array)
        for i in range(gap, a):
            graf(array)
            temp = array[i]
            j = i
            while j >= gap and array[j - gap] > temp:
                graf(array)
                array[j] = array[j - gap]
                j -= gap
            array[j] = temp
        gap //= 2


def nome(data):
    print('nome')
    i, j, size = 1, 2, len(data)
    while i < size:
        graf(data)
        if data[i - 1] <= data[i]:
            i, j = j, j + 1
        else:
            data[i - 1], data[i] = data[i], data[i - 1]
            i -= 1
            if i == 0:
                i, j = j, j + 1
    return data


def shaker_sort(arr):
    print('shaker')
    swapped = True
    start_index = 0
    end_index = len(arr) - 1
    while swapped:
        swapped = False
        for i in range(end_index):
            graf(arr)
            if arr[i] % 2:
                j = i + 1
                while arr[j] % 2 == 0 and j < end_index:
                    j += 1
                if arr[j] % 2:
                    if arr[i] < arr[j]:
                        arr[i], arr[j] = arr[j], arr[i]
                        swapped = True
        if not swapped:
            break
        swapped = False
        end_index = end_index - 1
        for i in range(end_index - 1, start_index, -1):
            graf(arr)
            if arr[i] % 2:
                j = i - 1
                while arr[j] % 2 == 0 and j > start_index:
                    j -= 1
                if arr[j] % 2:
                    if arr[i] > arr[j]:
                        arr[i], arr[j] = arr[j], arr[i]
                        swapped = True
        start_index = start_index + 1
    return arr


def graf(B):
    plt.clf()
    plt.plot(B)
    plt.draw()
    plt.gcf().canvas.flush_events()


stat_shellsort = time.time()
shellsort(A)
time_shellsort = time.time() - stat_shellsort  # 1.25

start_gnomesort = time.time()
nome(B)
time_gnomesort = time.time() - start_gnomesort  # 0.2970000000204891

start_shaker_sort = time.time()
shaker_sort(C)
time_shaker_sort = time.time() - start_shaker_sort  # 1151.936999999918
plt.ioff()

print(f'time_shellsort = {time_shellsort}\n'
      f'time_gnomesort = {time_gnomesort}\n'
      f'time_shaker_sort = {time_shaker_sort}')
